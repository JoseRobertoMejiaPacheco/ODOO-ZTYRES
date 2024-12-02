import io
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl import Workbook
from odoo.http import request, Response
from datetime import date
import calendar
import pandas as pd
from odoo import api, fields, models

class ExcelVentasPorVendedor(models.TransientModel):
    _name = 'excel_ventas_por_vendedor'
    
    def generate_excel_report(self, lista):
        # Generar el archivo Excel con openpyxl
        output = io.BytesIO()
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:  # Verificar si la hoja predeterminada existe
            wb.remove(wb['Sheet']) 

        # Definir el estilo del encabezado
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
#-------------------------------------------------------------------------------------------------------------
        for etiqueta, df in lista:
            sheet = wb.create_sheet(etiqueta)
            # Escribir los encabezados
            headers = df.columns.tolist()
            for col_num, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill

            # Escribir los datos
            for row_num, row in enumerate(df.itertuples(index=False, name=None), 2):
                for col_num, value in enumerate(row, 1):
                    sheet.cell(row=row_num, column=col_num, value=value)
#-------------------------------------------------------------------------------------------------------------
        # Ajustar el ancho de las columnas automáticamente
        for sheet in wb:
            for col in sheet.columns:
                max_length = 0
                col_letter = col[0].column_letter  # Obtener el nombre de la columna
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[col_letter].width = adjusted_width
        # Guardar el archivo Excel en memoria
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()