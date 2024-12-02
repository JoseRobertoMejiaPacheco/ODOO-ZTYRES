#-- coding: utf-8 --
from odoo import models, fields, api
from odoo import http
from odoo.http import request, Response
from openpyxl.worksheet.datavalidation import DataValidation
from werkzeug.urls import url_encode
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment,Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles.numbers import NumberFormat
from openpyxl import Workbook
class Invpromo(http.Controller):

    def calcular_porcentaje(self,monto_original, porcentaje, operacion):
        if operacion == "suma":
            resultado = monto_original + (monto_original * porcentaje / 100)
        elif operacion == "resta":
            resultado = monto_original - (monto_original * porcentaje / 100)
        else:
            raise ValueError("La operación debe ser 'suma' o 'resta'")
        return resultado

    def _set_data(self, partner_id,objects, ws, combo_financiero, currency_style,left_aligned_style,row, iva_percent=0):
        count = 0
        valid_values = objects.filtered(lambda item: item.volumen > 0 or item.promocion > 0 or item.outlet > 0)
        count_valid_values = len(valid_values) + 10
        ws['R8'] = f'=SUM(R10:R{count_valid_values})'
        for item in objects:
            if item.inventario_str not in ['No Disponible']:
                ws.cell(row=row, column=1).value = item.product_id.id or None or None
                ws.cell(row=row, column=1).style = left_aligned_style or None
                ws.cell(row=row, column=2).value = item.product_id.default_code or None
                ws.cell(row=row, column=2).style = left_aligned_style or None
                ws.cell(row=row, column=3).value = item.product_id.tire_measure_id.name or None
                ws.cell(row=row, column=4).value = item.product_id.layer_id.name or None
                ws.cell(row=row, column=5).value = item.product_id.speed_id.name or None
                ws.cell(row=row, column=6).value = item.product_id.index_of_load_id.name or None
                ws.cell(row=row, column=7).value = item.product_id.model_id.name or None
                ws.cell(row=row, column=8).value = item.product_id.brand_id.name or None
                ws.cell(row=row, column=9).value = item.product_id.tier_id.name or None
                ws.cell(row=row, column=10).value = item.inventario_str or None
                ws.cell(row=row, column=10).style = left_aligned_style or None
                ws.cell(row=row, column=11).value = item.product_id.product_dot_range or 'N/A' or None
                ws.cell(row=row, column=11).style = currency_style or None
                volumen = self.calcular_porcentaje(self.calcular_porcentaje(item.volumen,iva_percent,'suma'),partner_id.partner_discount_id.percent,'resta')
                ws.cell(row=row, column=12).value =  volumen if volumen > 0 else None
                ws.cell(row=row, column=12).style = currency_style                                
                ws.cell(row=row, column=13).value = item.promocion * (1 + iva_percent / 100) if item.promocion > 0 else None
                ws.cell(row=row, column=13).style = currency_style
                ws.cell(row=row, column=14).value = item.outlet * (1 + iva_percent / 100) if item.outlet > 0 else None
                ws.cell(row=row, column=14).style = currency_style                
                ws.cell(row=row, column=15).value = f'=IF(MIN(L{row}:N{row}) > 0, MIN(L{row}:N{row}), "")'
                ws.cell(row=row, column=15).style = currency_style
                if item.cantidad_gratis > 0:
                    ws.cell(row=row, column=16).value = f'=(O{row}*({item.condicion}-{item.cantidad_gratis})/{item.condicion})+2' 
                elif item.cantidad_gratis == 0:
                    ws.cell(row=row, column=16).value = item.promocion_paquetes * (1 + iva_percent / 100) or None
                ws.cell(row=row, column=16).style = currency_style
                if item.condicion > 0:
                    ws.cell(row=row, column=17).value = '%s Llantas' % (item.condicion) or None
                ws.cell(row=row, column=18).value = None                
                formula_pf = f'''=IF(R{row}<>"",IF(N{row}<>"",IF(R{row}>=IFERROR(VALUE(IFERROR(MID(Q{row},FIND(" ",Q{row})-2,2),MID(Q{row},FIND(" ",Q{row})-1,1))),""),"",""),IF(Q{row}<>"",IF(R{row}>=IFERROR(VALUE(IFERROR(MID(Q{row},FIND(" ",Q{row})-2,2),MID(Q{row},FIND(" ",Q{row})-1,1))),""),(P{row}*(1-IFERROR(VLOOKUP(SUM(R10:R{count_valid_values}),A6:B8,2,TRUE),"0%"))),IF(SUM(R10:R{count_valid_values})>=IFERROR(VLOOKUP(SUM(R10:R{count_valid_values}),A6:B8,1,TRUE),A6),(O{row}*(1-IFERROR(VLOOKUP(SUM(R10:R{count_valid_values}),A6:B8,2,TRUE),"0%"))),O{row})),IF(SUM(R10:R{count_valid_values})>=IFERROR(VLOOKUP(SUM(R10:R{count_valid_values}),A6:B8,1,TRUE),A6),(O{row}*(1-IFERROR(VLOOKUP(SUM(R10:R{count_valid_values}),A6:B8,2,TRUE),"0%"))),O{row}))),"")'''                                    
                formula_pf = formula_pf.format(iva_percent=iva_percent)
                ws.cell(row=row, column=19).value = formula_pf
                ws.cell(row=row, column=19).style = currency_style                
                ws.cell(row=row, column=20).value =  f'=IF(R{row}="","",IF(N{row}<>"",IF(R{row}>=IFERROR(VALUE(IFERROR(MID(Q{row},FIND(" ",Q{row})-2,2),MID(Q{row},FIND(" ",Q{row})-1,1))),""),P{row},N{row}),S{row}))'
                ws.cell(row=row, column=20).style = currency_style
                ws.cell(row=row, column=21).value = f'=IFERROR(T{row} * (1-{combo_financiero}), "")'
                ws.cell(row=row, column=21).style = currency_style
                row += 1
        return row - 1

    def _set_combo_financiero(self,partner_id,ws):
        tupla_valores = partner_id.get_row_values_financiero()
        tupla_valores.append((0, 0))
        operaciones_financiero = list(map(lambda x: f"{x[1]}%", tupla_valores))
        dv_financiero = DataValidation(type="list", formula1='"' + ','.join(operaciones_financiero) + '"', allow_blank=True)
        combo_financiero = 'U9'
        dv_financiero.add(ws[combo_financiero])
        ws.add_data_validation(dv_financiero)
        ws[combo_financiero].font = Font(bold=False)
        return combo_financiero

    def get_left_aligned_style(self):
        currency_style = NamedStyle(name='align')
        currency_style.font = Font(bold=False)
        currency_style.border = Border(
            left=Side(border_style=None),
            right=Side(border_style=None),
            top=Side(border_style=None),
            bottom=Side(border_style=None)
        )
        currency_style.alignment = Alignment(horizontal='left', vertical='center')        
        return currency_style
      
    def _get_currency_style(self):        
        currency_style = NamedStyle(name='Currency')
        currency_style.font = Font(bold=False)
        currency_style.border = Border(left=Side(border_style=None),
                                    right=Side(border_style=None),
                                    top=Side(border_style=None),
                                    bottom=Side(border_style=None))
        currency_style.alignment = Alignment(horizontal='right', vertical='center')
        currency_style.number_format = ' #,##0.00'  
        return currency_style
    
    def setup_work_sheet_columns(self,ws,res_partner):
        ws.merge_cells('A1:B1')
        ws.merge_cells('C1:D1')
        ws.merge_cells('A2:D2')
        ws.merge_cells('A3:D3')
        ws.merge_cells('A4:D4')
        ws['C1'] = res_partner
        ws['A1'] = 'Lista Marzo 2024'
        ws['A2'] = 'Inventario + promociones Marzo con IVA'
        ws['A3'] = 'bases de la promoción paquetes'
        ws['A4'] = 'Precios outlet solo aplica descuentos financieros'          
        ws['R9'] = 'Descuento Logístico'
        ws['U8'] = 'Financieros'                
        ws['A1'].fill = PatternFill(start_color='1a1818', end_color='1a1818', fill_type='solid')                  
        ws['A2'].fill = PatternFill(start_color='1a1818', end_color='1a1818', fill_type='solid')  
        ws.merge_cells('L8:P8')
        ws['L8'] = 'LISTA DE PRECIOS Y PROMOCIONES'
        ws['L8'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        ws['L8'].fill = PatternFill(start_color='1a1818', end_color='1a1818', fill_type='solid')
        ws.merge_cells('A5:B5')
        ws['A5'] = 'Descuentos Logísticos'       
        ws['A5'].fill = PatternFill(start_color='1a1818', end_color='1a1818', fill_type='solid')
      
        
    @http.route('/inv_promo/inv_promo/objects', auth='public')
    def list(self, **kw):
        objects = request.env['inv_promo.report'].sudo().insert_data()
        res_partner = request.env['res.partner'].sudo().browse(7053)        
        wb = Workbook()
        ws = wb.active
        self.setup_work_sheet_columns(ws,res_partner)
        row = 6
        for cantidad, descuento in res_partner.get_row_values_logistico():
            ws.cell(row=row, column=1).value = cantidad
            ws.cell(row=row, column=2).value = '%s%%' % (str(descuento))
            row += 1
        row = 3
        # for cantidad, descuento in res_partner.get_row_values_financiero():
        #     ws.cell(row=row, column=9).value = cantidad
        #     ws.cell(row=row, column=10).value = '%s%%' % (str(descuento))
        #     row += 1
        row = 9
        headers = [
    ('id',.001),
    ('Código', 4),
    ('Medida', 5.3),
    ('Capas',.001),
    ('Vel',.001),
    ('Carga', 3),
    ('Modelo', 15),
    ('Marca', 8),
    ('Tier',.001),
    ('Inv',.001),
    ('DOT',2),
    ('Volumen', 3),
    ('Promoción',3),
    ('Outlet',3),
    ('Mejor Condición',3),
    ('Paquetes',3),
    ('Condición',3),
    ('Pedido', 2),
    ('Logístico',3),
    ('Precio Final',3),
    ('0%',3),
        ]
        for col, (header,max_len) in enumerate(headers, start=1):
            max_length = 0
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=False)
            wrap_alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            cell.alignment = wrap_alignment
            col_letra = ws.cell(row=row, column=col).column_letter
            adjusted_width = (max_len+5) * 1.4       
            ws.column_dimensions[col_letra].width = adjusted_width


        ws.freeze_panes = 'A10'
        row += 1
        currency_style = self._get_currency_style()
        align_style = self.get_left_aligned_style()
        combo_financiero = self._set_combo_financiero(res_partner,ws)        
        count_valid_values = self._set_data(res_partner,objects,ws,combo_financiero,currency_style,align_style,row,0)    
            
        ws_copia = wb.copy_worksheet(ws)
        self._set_data(res_partner,objects,ws_copia,combo_financiero,currency_style,align_style,row,16)
        ws_copia.title = 'Precios Con Iva'
        ws_copia.freeze_panes = 'A10'
        
        ws.title = 'Precios Sin Iva'
        
        # Definir la referencia para la primera tabla en la hoja de trabajo principal
        tab_ref = f'A9:U{count_valid_values}'  # Cambia esto según tus necesidades
        tab = Table(displayName='Table1', ref=tab_ref)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=False, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)
        
        # Definir la referencia para la segunda tabla en la hoja de trabajo de copia
        tab2 = Table(displayName="Table2", ref=tab_ref)
        style2 = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                showLastColumn=False, showRowStripes=False, showColumnStripes=False)
        tab2.tableStyleInfo = style2
        ws_copia.add_table(tab2)
        sheets = wb.sheetnames
        new_font = Font(name='Arial', size=11,bold=False)
        for sheet_name in sheets:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    cell.font = new_font   
                            
        ws['F1'].font = Font(color='FFFFFF', bold=False)
        ws['I1'].font = Font(color='FFFFFF', bold=False)        
        ws['L8'].font = Font(color='FFFFFF', bold=False)   
        ws['A1'].font = Font(color='FFFFFF', bold=False)
        ws['T8'].font = Font(color='1a1818', bold=False)
        ws['A2'].font = Font(color='FFFFFF', bold=False) 
        ws['A5'].font = Font(color='FFFFFF', bold=False) 
        ws['I2'].font = Font(bold=False)
        ws['J2'].font = Font(bold=False)          
        ws['F2'].font = Font(bold=False)
        ws['G2'].font = Font(bold=False)  
        
        ws_copia['F1'].font = Font(color='FFFFFF', bold=False)
        ws_copia['I1'].font = Font(color='FFFFFF', bold=False)        
        ws_copia['L8'].font = Font(color='FFFFFF', bold=False)   
        ws_copia['A1'].font = Font(color='FFFFFF', bold=False)
        ws_copia['T8'].font = Font(color='1a1818', bold=False)
        ws_copia['A2'].font = Font(color='FFFFFF', bold=False) 
        ws_copia['A5'].font = Font(color='FFFFFF', bold=False) 
        ws_copia['I2'].font = Font(bold=False)
        ws_copia['J2'].font = Font(bold=False)          
        ws_copia['F2'].font = Font(bold=False)
        ws_copia['G2'].font = Font(bold=False)           

        new_font = Font(name='Arial', size=11, bold=False)
        white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')  
        for sheet_name in sheets:
            sheet = wb[sheet_name]
            tab_ref = f'A9:U{count_valid_values}'
            cell_range = sheet[tab_ref]
            for row in cell_range:
                for cell in row:
                    cell.font = new_font
                    cell.fill = white_fill     
            sheet.sheet_view.zoomScale = 93   

        color_gris_claro = "bab6b6"

        rango_celdas = ws['A9:K9']

        for fila in rango_celdas:
            for celda in fila:
                celda.fill = PatternFill(start_color=color_gris_claro, end_color=color_gris_claro, fill_type="solid")

        color_gris_claro = "edede8"
        rango_celdas = ws['K9:Q9']
        for fila in rango_celdas:
            for celda in fila:
                celda.fill = PatternFill(start_color=color_gris_claro, end_color=color_gris_claro, fill_type="solid")

        color_gris_claro = "faea57"
        rango_celdas = ws['Q9:U9']
        for fila in rango_celdas:
            for celda in fila:
                celda.fill = PatternFill(start_color=color_gris_claro, end_color=color_gris_claro, fill_type="solid")                                                 

###
        color_gris_claro = "bab6b6"

        rango_celdas = ws_copia['A9:K9']

        for fila in rango_celdas:
            for celda in fila:
                celda.fill = PatternFill(start_color=color_gris_claro, end_color=color_gris_claro, fill_type="solid")

        color_gris_claro = "edede8"
        rango_celdas = ws_copia['K9:Q9']
        for fila in rango_celdas:
            for celda in fila:
                celda.fill = PatternFill(start_color=color_gris_claro, end_color=color_gris_claro, fill_type="solid")

        color_gris_claro = "faea57"
        rango_celdas = ws_copia['Q9:U9']
        for fila in rango_celdas:
            for celda in fila:
                celda.fill = PatternFill(start_color=color_gris_claro, end_color=color_gris_claro, fill_type="solid")                                                 
                
        wb.save('/mnt/extra-addons/Lista de Precios%s.xlsx' % (res_partner.name or ' M'))

