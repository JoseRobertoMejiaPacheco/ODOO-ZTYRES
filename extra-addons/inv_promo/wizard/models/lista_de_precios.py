from odoo import _, api, fields, models
from odoo import models, fields, api
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from odoo.exceptions import UserError,ValidationError
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from pytz import timezone
import openpyxl
import base64
import io

codes=[22094,22075,22356,51533,22305,22088,22518,22125,22589,22804,49875,51810,57249,22524,51129,48697,51575,22711,51715,22137,22712,22603,
       22822,22605,58459,22785,22155,48878,22835,22816,22158,22161,22823,22814,50301,22325,22260,22175,58485,48952,22510,50199,48953,22604,
       22625,22422,22776,22560,22828,22170,22652,22811,22633,22777,48954,22817,51855,22288,22284,22295,22484,22299,22833,22516,22281,22826,
       22355,22301,58463,49622,58464,51857,22827,57226,22393,22808,22252,22419
]

codes2=[58907,58908,58909,58910,58911,58912,58913,58914,58915,58917,58916,58918,58919
]

codes3=[52645,52646,2448,12060,12050,12057,2196,52649,52650,12058,57528,12116,52651,2290,11985,2299,2303,2197,11981,58384,58479,12086,4269,
        2364,2235,49826,2451,2292,2293,2368,12022,4432,12027,2274,12039,51460,12046,52665,52723,4433,52674,12040,49831,52676,2289,52677,52678,
        2295,4383,49793,52682,58484,12042,12128,49794,12105,57523,12129,12001,4396,4358,4384,12104,12004,4382,2264,12005,4335,12099,12103,4373,
        12018,4366,12101,12012,12108,4394,12014,12111,12019,12008,2410,50154,2351,2323,52010,2402,2405,2228,2277,2195,2455,49843,2325,49852,
        49848,2206,49825,49849,2414,2467,2438,2211,2417,2385,2428,2422,2241,2372,2384,2407,57265,57255,57504,52056,57275,2314,57507,57283,2399,
        57259,57240,4253,4285,4164,4289,4275,4403,52714,4256,4187,4435,51979,4401,4237,4188,4197,4176,4193,52687,52725,49797,52735,12009,12085,
        12023,12137,12043,12065,12125,12070,12003,12067
]
       
WHITE = 'FFFFFF'
BLACK = '1A1818'
YELLOW = 'FFFF00'
YELLOW_2 = 'FFC000'
DARK_GREEN = '19B050'
LIGTH_GREEN = '92D050'
RED = 'FF0000'
LIGTH_GRAY = 'F2F2F2'
DARK_GRAY = '404040'
BLUE = '95B3D7'

class ListaDePrecios(models.TransientModel):
    _name = 'inv_promo.lista_precios_wizard'
    _description = 'Lista de Precios'
    
    partner_id = fields.Many2one('res.partner', string='Cliente')
    volume_profile= fields.Many2one('discount_profiles.volume.discount', string='Volumen')
    financial_profile = fields.Many2one('discount_profiles.financial.discount', string='Financiero')
    logistic_profile = fields.Many2one('discount_profiles.logistic.discount', string='Logístico')
    file_data = fields.Binary('File')

    def get_profile_data(self,partner_id=False):
        if partner_id:    
            financial = partner_id.get_row_values_financiero()
            logistic = partner_id.get_row_values_logistico()
        else:
            financial = self.partner_id.get_row_values_financiero(self.financial_profile)
            logistic = self.partner_id.get_row_values_logistico(self.logistic_profile)    
        
        # Convertir valores numéricos a cadenas de porcentaje y agregar '0%'
        financial_percentages = ['0%'] + [f"{tupla[1]}%" for tupla in financial]
        logistic_percentages = ['0%'] + [f"{tupla[1]}%" for tupla in logistic]
        
        # Ordenar las listas de porcentajes
        financial_percentages.sort()
        logistic_percentages.sort()
        
        return financial_percentages, logistic_percentages
########################################################################## Efervecente#########################################################################################################################################################################################
    def _get_metadata(self,res_partner):
        data = ''
        tz = timezone('America/Mexico_City')  # Definir la zona horaria de México
        if res_partner:
            # Convertir la fecha de creación a la zona horaria de México
            created_date = fields.Datetime.to_string(fields.Datetime.context_timestamp(self, fields.Datetime.from_string(self.create_date or fields.Datetime.now())).astimezone(tz))
            data = f'Creado por: {self.env.user.name}, Fecha de creación: {created_date}, Cliente: {res_partner.name}, Volumen: {res_partner.volume_profile.name}, Crédito: {res_partner.financial_profile .name}, Logístico: {res_partner.logistic_profile .name}'
        else:
            created_date = fields.Datetime.to_string(fields.Datetime.context_timestamp(self, fields.Datetime.from_string(self.create_date)).astimezone(tz))
            data = f'Creado por: {self.env.user.name}, Fecha de creación: {created_date}, Volumen: {self.volume_profile.name}, Crédito: {self.financial_profile .name}, Logístico: {self.logistic_profile .name}'            
        return data
        
    def download_report(self):
        if self.partner_id:
            return self.get_xlsx_report(self.partner_id)
        elif not self.partner_id and self.volume_profile and self.financial_profile  and self.logistic_profile :
            return self.get_xlsx_report(self.partner_id)

    def calcular_porcentaje(self,monto_original, porcentaje, operacion):
        if operacion == "suma":
            resultado = monto_original + (monto_original * porcentaje / 100)
        elif operacion == "resta":
            resultado = monto_original - (monto_original * porcentaje / 100)
        else:
            raise ValueError("La operación debe ser 'suma' o 'resta'")
        return resultado

    def insert_table_sheet1_3(self, sheet, table_data, table_name):
        headers = list(table_data[0].keys()) if table_data else []
        num_rows = len(table_data)
        if num_rows == 0 or not headers:
            return  # No hay datos o encabezados, salir sin crear la tabla

        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=14, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)

        for row_idx, row_data in enumerate(table_data, start=15):
            for col_idx, header in enumerate(headers, start=1):
                if header == "Mejor Condición":
                    sheet.cell(row=row_idx, column=col_idx).value = f'=IF(IF(MIN(Q{row_idx}:S{row_idx}) > 0, MIN(Q{row_idx}:S{row_idx}), "") = S{row_idx}, S{row_idx}, IF(MIN(Q{row_idx}:S{row_idx}) > 0, MIN(Q{row_idx}:S{row_idx}), "") * (1-IF(ISNUMBER($T$13), $T$13, 0)))'                                     
                elif header == "Descuento Potencial":
                    sheet.cell(row=row_idx, column=col_idx).value = f'=IF($W$11 = "", "", (T{row_idx} * (1 - $U$13)))'
                elif header == "Financiero":
                    sheet.cell(row=row_idx, column=col_idx).value = f'=IF($W$11 <> "", "",IF(U{row_idx} = "", (T{row_idx} * (1 - $V$13)), (U{row_idx} * (1 - $V$13))))'                               
                elif header == "Total":
                    sheet.cell(row=row_idx, column=col_idx).value = f'=IF(W{row_idx} = "", "", (T{row_idx} * W{row_idx}))'
                else:
                     sheet.cell(row=row_idx, column=col_idx).value = row_data.get(header, "")

        end_col_letter = get_column_letter(len(headers))
        table_ref = f"A14:{end_col_letter}{num_rows + 14}"
        table = Table(displayName=table_name, ref=table_ref)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        table.tableStyleInfo = style
        sheet.add_table(table)
        return num_rows + 14, len(headers)  # Retorna el número de filas y columnas
####################################################################################################
    def insert_table_sheet4(self, sheet, table_data, table_name):
            headers = list(table_data[0].keys()) if table_data else []
            num_rows = len(table_data)
            if num_rows == 0 or not headers:
                return  # No hay datos o encabezados, salir sin crear la tabla

            for col_idx, header in enumerate(headers, start=1):
                cell = sheet.cell(row=14, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True)

            for row_idx, row_data in enumerate(table_data, start=15):
                for col_idx, header in enumerate(headers, start=1):
                    if header == "300":
                        sheet.cell(row=row_idx, column=col_idx).value = f'=T{row_idx} * (1 - $V$12)'
                    elif header == "600":
                        sheet.cell(row=row_idx, column=col_idx).value = f'=T{row_idx} * (1 - $W$12)'
                    elif header == "1000":
                        sheet.cell(row=row_idx, column=col_idx).value = f'=T{row_idx} * (1 - $X$12)'
                    elif header == "Total":
                        sheet.cell(row=row_idx, column=col_idx).value = f'=IF(Y{row_idx} = "", "", (Y{row_idx} * T{row_idx}))'
                    elif header == "Mejor Condición":
                        sheet.cell(row=row_idx, column=col_idx).value = f'=IF(IF(MIN(Q{row_idx}:S{row_idx}) > 0, MIN(Q{row_idx}:S{row_idx}), "") = S{row_idx}, S{row_idx}, IF(MIN(Q{row_idx}:S{row_idx}) > 0, MIN(Q{row_idx}:S{row_idx}), "") * (1-IF(ISNUMBER($T$13), $T$13, 0)))'                                     
                    elif header == "Descuento Potencial":
                        sheet.cell(row=row_idx, column=col_idx).value = f'=IF($W$11 = "", "", (T{row_idx} * (1 - $U$13)))'   
                    else:
                        sheet.cell(row=row_idx, column=col_idx).value = row_data.get(header, "")

            end_col_letter = get_column_letter(len(headers))
            table_ref = f"A14:{end_col_letter}{num_rows + 14}"
            table = Table(displayName=table_name, ref=table_ref)
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            table.tableStyleInfo = style
            sheet.add_table(table)
            return num_rows + 14, len(headers)  # Retorna el número de filas y columnas
####################################################################################################
    def insert_table_sheet5(self, sheet, table_data, table_name):
            headers = list(table_data[0].keys()) if table_data else []
            num_rows = len(table_data)
            if num_rows == 0 or not headers:
                return  # No hay datos o encabezados, salir sin crear la tabla

            for col_idx, header in enumerate(headers, start=1):
                cell = sheet.cell(row=14, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True)

            for row_idx, row_data in enumerate(table_data, start=15):
                for col_idx, header in enumerate(headers, start=1):
                    if header == "60":
                        sheet.cell(row=row_idx, column=col_idx).value = f'=T{row_idx} * (1 - $V$12)'
                    elif header == "150":
                        sheet.cell(row=row_idx, column=col_idx).value = f'=T{row_idx} * (1 - $W$12)'
                    elif header == "350":
                        sheet.cell(row=row_idx, column=col_idx).value = f'=T{row_idx} * (1 - $X$12)'
                    elif header == "Total":
                        sheet.cell(row=row_idx, column=col_idx).value = f'=IF(Y{row_idx} = "", "", (Y{row_idx} * T{row_idx}))'
                    elif header == "Mejor Condición":
                        sheet.cell(row=row_idx, column=col_idx).value = f'=IF(IF(MIN(Q{row_idx}:S{row_idx}) > 0, MIN(Q{row_idx}:S{row_idx}), "") = S{row_idx}, S{row_idx}, IF(MIN(Q{row_idx}:S{row_idx}) > 0, MIN(Q{row_idx}:S{row_idx}), "") * (1-IF(ISNUMBER($T$13), $T$13, 0)))'                                     
                    elif header == "Descuento Potencial":
                        sheet.cell(row=row_idx, column=col_idx).value = f'=IF($W$11 = "", "", (T{row_idx} * (1 - $U$13)))'   
                    else:
                        sheet.cell(row=row_idx, column=col_idx).value = row_data.get(header, "")

            end_col_letter = get_column_letter(len(headers))
            table_ref = f"A14:{end_col_letter}{num_rows + 14}"
            table = Table(displayName=table_name, ref=table_ref)
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            table.tableStyleInfo = style
            sheet.add_table(table)
            return num_rows + 14, len(headers)  # Retorna el número de filas y columnas
####################################################################################################
    def set_combos(self,sheet, combo_data):
        combo_cells = []
        for data in combo_data:
            cell_ref = data['cell_ref']
            values = data['values']
            combo_cells.append(self._set_combo(sheet, cell_ref, values))
        return combo_cells

    def _set_combo(self,sheet, cell_ref, values):
        dv = DataValidation(type="list", formula1='"' + ','.join(values) + '"', allow_blank=True)
        sheet.add_data_validation(dv)
        dv.add(sheet[cell_ref])
        sheet[cell_ref].font = Font(bold=False)
        return cell_ref

    def apply_styles(self,sheet, cell_range, font_style, align_style, fill_style, border_style, num_format):
        for row in cell_range:
            for cell in row:
                cell.font = font_style
                cell.alignment = align_style
                cell.fill = fill_style
                cell.border = border_style
                if num_format:
                    cell.number_format = num_format

    def format_column(self, sheet, column_name, data_type):
        col_idx = column_index_from_string(column_name)
        num_format = None        
        if data_type == 'date':
            num_format = 'DD/MM/YYYY'
        elif data_type == 'currency':
            num_format = '$#,##0.00'
        elif data_type == 'number':
            num_format = '#,##0'
        elif data_type == 'string':
            num_format = None  # No specific format
        for row in sheet.iter_rows(min_col=col_idx, max_col=col_idx, min_row=15):
            for cell in row:
                if cell.value == 0:
                    cell.value = None  # Set cell value to None to leave it empty
                elif num_format:
                    cell.number_format = num_format

    def set_white_fill(self, sheet, start_cell, num_rows, num_cols):
        start_col_letter = ''.join(filter(str.isalpha, start_cell))
        start_row_number = ''.join(filter(str.isdigit, start_cell))
        
        start_col_idx = column_index_from_string(start_col_letter)
        end_col_idx = start_col_idx + num_cols -1
        end_row_number = int(start_row_number) + num_rows

        end_col_letter = get_column_letter(end_col_idx)
        cell_range = f'{start_col_letter}{start_row_number}:{end_col_letter}{end_row_number}'

        fill_style = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
        alignment_style = Alignment(horizontal='center', vertical='center')
        for row in sheet[cell_range]:
            for cell in row:
                cell.fill = fill_style
                cell.alignment = alignment_style

    def format_cell(self,sheet, cell_ref, font_name, font_size, font_color, bold, fill_color, border, align, top_align, wrap_text=False, num_format=None, value=None, data_type='string'):
        # Parse the cell reference or range
        if ':' in cell_ref:
            start_ref, end_ref = cell_ref.split(':')
            cell_range = sheet[start_ref:end_ref]
            merged_cells = True
        else:
            start_ref = cell_ref
            cell_range = [[sheet[start_ref]]]
            merged_cells = False

        # Define styles
        font_style = Font(name=font_name, size=font_size, color=font_color, bold=bold)
        align_style = Alignment(horizontal=align, vertical=top_align, wrap_text=wrap_text)
        fill_style = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        border_style = Border(left=Side(border_style=border),
                            right=Side(border_style=border),
                            top=Side(border_style=border),
                            bottom=Side(border_style=border))

        # Set number format if applicable
        if data_type == 'date':
            num_format = 'yyyy-mm-dd'
        elif data_type == 'currency':
            num_format = '$#,##0.00'
        elif num_format is None:
            num_format = '#,##0.00' if data_type == 'number' else None
            
        # Apply styles
        self.apply_styles(sheet, cell_range, font_style, align_style, fill_style, border_style, num_format)

        # Set value to the first cell
        if value is not None:
            sheet[start_ref].value = value

        # Adjust column width based on content length
        start_col, start_row = start_ref[0], int(start_ref[1:])
        end_col = end_ref[0] if merged_cells else start_col
        end_row = int(end_ref[1:]) if merged_cells else start_row

        for col in range(ord(start_col), ord(end_col) + 1):
            col_letter = get_column_letter(col)
            max_len = 0
            for row in range(start_row, end_row + 1):
                cell = sheet[col_letter + str(row)]
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            adjusted_width = (max_len + 5) * 1.4
            sheet.column_dimensions[col_letter].width = adjusted_width

        # Merge cells if needed
        if merged_cells:
            sheet.merge_cells(start_row=start_row, start_column=ord(start_col) - 64,
                            end_row=end_row, end_column=ord(end_col) - 64)

    def create_attachment(self,res_id,partner_id):
        encoded_data = self.get_xlsx_report(partner_id,return_base_64=True)
        attachment = self.env['ir.attachment'].create({
            'name': f'{partner_id.name or "Lista de Precios"}.xlsx',
            'type': 'binary',
            'datas': encoded_data,
            'store_fname': f'{partner_id.name or "Lista de Precios"}.xlsx',
            'mimetype': 'application/vnd.ms-excel',
            'res_model': self._name,
            'res_id': res_id or partner_id.id,
        })
        return attachment 

    def color_cells_based_on_condition(self, sheet):
        outlet_col_idx = column_index_from_string('S')  # Replace 'A' with actual column letter for 'Outlet'
        mejor_condicion_col_idx = column_index_from_string('T')  # Replace 'B' with actual column letter for 'Mejor Condición'
        ids_in_codes_brig = column_index_from_string('A')
        # Define el rango de celdas que deseas pintar
        start_column = 'A'
        end_column = 'S'

        # Itera sobre las celdas en el rango y aplica el relleno si el valor está en codes
        for row in sheet.iter_rows(min_row=15, min_col=22, max_col=22):
            for cell in row:
                cell.fill = PatternFill(start_color='ECF0DE', end_color='ECF0DE', fill_type='solid')
                cell.font = Font(color='526225')
        
        for row in sheet.iter_rows(min_row=15):  # Start from row 2 to skip header
            outlet_cell = row[outlet_col_idx - 1]  # Adjust index to zero-based
            mejor_condicion_cell = row[mejor_condicion_col_idx - 1]  # Adjust index to zero-based
            # Example condition: Color 'Outlet' column yellow if cell value is not empty
            if isinstance(outlet_cell.value, (int, float)) and outlet_cell.value > 0:
                outlet_cell.fill = PatternFill(start_color='FDFFCD', end_color='FDFFCD', fill_type='solid')
                mejor_condicion_cell.fill = PatternFill(start_color='FDFFCD', end_color='FDFFCD', fill_type='solid')
                
        # Itera sobre las celdas en el rango y aplica el relleno si el valor está en codes
        for row in sheet.iter_rows(min_row=15):
            id_cell = row[ids_in_codes_brig - 1]
            for col in range(ord(start_column), ord(end_column) + 1):
                if id_cell.value in codes:
                    cell = row[col - ord(start_column)]
                    cell.fill = PatternFill(start_color='E6B8B8', end_color='E6B8B8', fill_type='solid')
                    
        for row in sheet.iter_rows(min_row=15):
            id_cell = row[ids_in_codes_brig - 1]
            for col in range(ord(start_column), ord(end_column) + 1):
                if id_cell.value in codes2:
                    cell = row[col - ord(start_column)]
                    cell.fill = PatternFill(start_color='DDE5F2', end_color='DDE5F2', fill_type='solid')
                    
        for row in sheet.iter_rows(min_row=15):
            id_cell = row[ids_in_codes_brig - 1]
            for col in range(ord(start_column), ord(end_column) + 1):
                if id_cell.value in codes3:
                    cell = row[col - ord(start_column)]
                    cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
   
    def color_cells_based_on_condition_sheet3(self, sheet):
        outlet_col_idx = column_index_from_string('S')  # Replace 'A' with actual column letter for 'Outlet'
        mejor_condicion_col_idx = column_index_from_string('T')  # Replace 'B' with actual column letter for 'Mejor Condición'

        for row in sheet.iter_rows(min_row=15):  # Start from row 2 to skip header
            outlet_cell = row[outlet_col_idx - 1]  # Adjust index to zero-based
            mejor_condicion_cell = row[mejor_condicion_col_idx - 1]  # Adjust index to zero-based
            # Example condition: Color 'Outlet' column yellow if cell value is not empty
            if isinstance(outlet_cell.value, (int, float)) and outlet_cell.value > 0:
                outlet_cell.fill = PatternFill(start_color='FDFFCD', end_color='FDFFCD', fill_type='solid')
                mejor_condicion_cell.fill = PatternFill(start_color='FDFFCD', end_color='FDFFCD', fill_type='solid')
                
    def color_cells_based_on_condition_sheet4(self, sheet):
        outlet_col_idx = column_index_from_string('S')  # Replace 'A' with actual column letter for 'Outlet'
        mejor_condicion_col_idx = column_index_from_string('T')  # Replace 'B' with actual column letter for 'Mejor Condición'
        
        for row in sheet.iter_rows(min_row=15):  # Start from row 2 to skip header
            outlet_cell = row[outlet_col_idx - 1]  # Adjust index to zero-based
            mejor_condicion_cell = row[mejor_condicion_col_idx - 1]  # Adjust index to zero-based
            # Example condition: Color 'Outlet' column yellow if cell value is not empty
            if isinstance(outlet_cell.value, (int, float)) and outlet_cell.value > 0:
                outlet_cell.fill = PatternFill(start_color='FDFFCD', end_color='FDFFCD', fill_type='solid')
                mejor_condicion_cell.fill = PatternFill(start_color='FDFFCD', end_color='FDFFCD', fill_type='solid')
                
        for row in sheet.iter_rows(min_row=15, min_col=22, max_col=22):
            for cell in row:
                cell.fill = PatternFill(start_color='C5D9F1', end_color='C5D9F1', fill_type='solid')
                cell.font = Font(color='538DD5')
        for row in sheet.iter_rows(min_row=15, min_col=23, max_col=23):
            for cell in row:
                cell.fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
                cell.font = Font(color='16365C')
        for row in sheet.iter_rows(min_row=15, min_col=24, max_col=24):
            for cell in row:
                cell.fill = PatternFill(start_color='C5D9F1', end_color='C5D9F1', fill_type='solid')
                cell.font = Font(color='538DD5')

    def insert_formula(self,sheet, column, start_row, template):
        current_row = start_row
        while current_row <= sheet.max_row:
            cell = f"{column}{current_row}"
            formula = template.replace('{column}', column).replace('{row}', str(current_row))
            sheet[cell].value = formula
            current_row += 1
            print(current_row)

    def get_dict_data(self, objects, sheet_name,partner_id=False):
        if sheet_name == 'Precios Con Iva':
            c = self.env['additional_discounts.products_line'].search([('additional_prod_id.active','=',True)]).mapped('product_tmpl_id').ids + objects.ids
            ids_tuple = tuple(c)
            
            # Construir la consulta SQL con los IDs en la cláusula IN
            query = """
                SELECT id 
                FROM inv_promo_report 
                WHERE id IN %s
            """
            # Ejecutar la consulta con la tupla de IDs
            self.env.cr.execute(query, (ids_tuple,))
            result = self.env.cr.fetchall()
            
            ids = [row[0] for row in result]
            
            objects = objects.browse(ids)

        if sheet_name == 'Onyx':
            
            codes_tuple = tuple(codes2)
            
            ids_tuple = tuple(objects.ids)
            
            query = """
                SELECT id 
                FROM inv_promo_report 
                WHERE product_id IN %s AND id IN %s
            """
            self.env.cr.execute(query, (codes_tuple, ids_tuple))
            result = self.env.cr.fetchall()
            
            ids = [row[0] for row in result]
            
            objects = objects.browse(ids)
            data_list = []     
            for obj in objects:
                data_dict = {                
                    'id': obj.product_id.id,            
                    'Código': obj.default_code,
                    'Medida': obj.tire_measure_id.name if obj.tire_measure_id else None,
                    'Capas': obj.layer_id.name if obj.layer_id else None,
                    'Vel': obj.speed_id.name if obj.speed_id else None,
                    'Carga': obj.index_of_load_id.name if obj.index_of_load_id else None,
                    'Modelo': obj.model_id.name if obj.model_id else None,
                    'Marca': obj.brand_id.name if obj.brand_id else None,
                    'Tipo': obj.product_id.type_id.name or None,
                    'Seg': obj.product_id.segment_id.name or None,
                    'Tier': obj.tier_id.name if obj.tier_id else None,
                    'Inv.': obj.inventario_str,
                    'Cantidad Disponible': obj.available,
                    'Trans.': (obj.transito_str + obj.backorder_str) or None,
                    'Arribo': obj.fecha_aprox or None,
                    'DOT': obj.product_id.product_dot_range or 'N/A' or None,  
                    }
                if self.partner_id:
                    data_dict.update({'Volumen': self.calcular_porcentaje(self.calcular_porcentaje(obj.volumen, 16, 'suma'), self.partner_id.volume_profile.percent, 'resta'),})
                else:
                    data_dict.update({'Volumen': self.calcular_porcentaje(self.calcular_porcentaje(obj.volumen, 16, 'suma'), self.volume_profile.percent, 'resta'),})
                data_list.append(data_dict)            
                data_dict.update({
                    'Promoción': obj.promocion * 1.16,
                    'Outlet': obj.outlet * 1.16,
                    'Mejor Condición': "",
                    'Descuento Potencial': "",
                    '300': "",
                    '600': "",
                    '1000': "",
                    'Pedido': "",
                    'Total': "",
                })
            return data_list
        
        if sheet_name == 'Tier 4':
            # Convertir la lista de códigos en una tupla para SQL
            codes_tuple = tuple(codes3)
            
            # Obtener los IDs de los objetos y convertirlos en una tupla para SQL
            ids_tuple = tuple(objects.ids)
            
            # Construir la consulta SQL con los códigos en la cláusula NOT IN y los IDs en la cláusula IN
            query = """
                SELECT id 
                FROM inv_promo_report 
                WHERE product_id IN %s AND id IN %s
            """
            
            self.env.cr.execute(query, (codes_tuple, ids_tuple))
            result = self.env.cr.fetchall()
            
            ids = [row[0] for row in result]
            
            objects = objects.browse(ids)
            data_list = []     
            for obj in objects:
                data_dict = {                
                    'id': obj.product_id.id,            
                    'Código': obj.default_code,
                    'Medida': obj.tire_measure_id.name if obj.tire_measure_id else None,
                    'Capas': obj.layer_id.name if obj.layer_id else None,
                    'Vel': obj.speed_id.name if obj.speed_id else None,
                    'Carga': obj.index_of_load_id.name if obj.index_of_load_id else None,
                    'Modelo': obj.model_id.name if obj.model_id else None,
                    'Marca': obj.brand_id.name if obj.brand_id else None,
                    'Tipo': obj.product_id.type_id.name or None,
                    'Seg': obj.product_id.segment_id.name or None,
                    'Tier': obj.tier_id.name if obj.tier_id else None,
                    'Inv.': obj.inventario_str,
                    'Cantidad Disponible': obj.available,
                    'Trans.': (obj.transito_str + obj.backorder_str) or None,
                    'Arribo': obj.fecha_aprox or None,
                    'DOT': obj.product_id.product_dot_range or 'N/A' or None,  
                    }
                if self.partner_id:
                    data_dict.update({'Volumen': self.calcular_porcentaje(self.calcular_porcentaje(obj.volumen, 16, 'suma'), self.partner_id.volume_profile.percent, 'resta'),})
                else:
                    data_dict.update({'Volumen': self.calcular_porcentaje(self.calcular_porcentaje(obj.volumen, 16, 'suma'), self.volume_profile.percent, 'resta'),})
                data_list.append(data_dict)            
                data_dict.update({
                        'Promoción': obj.promocion * 1.16,
                        'Outlet': obj.outlet * 1.16,
                        'Mejor Condición': "",
                        'Descuento Potencial': "",
                        '60': "",
                        '150': "",
                        '350': "",
                        'Pedido': "",
                        'Total': "",
                    })
            return data_list
        
        if sheet_name == 'P. BRIDGESTONE':
            # Convertir la lista de códigos en una tupla para SQL
            codes_tuple = tuple(codes)
            
            # Obtener los IDs de los objetos y convertirlos en una tupla para SQL
            ids_tuple = tuple(objects.ids)
            
            # Construir la consulta SQL con los códigos en la cláusula NOT IN y los IDs en la cláusula IN
            query = """
                SELECT id 
                FROM inv_promo_report 
                WHERE product_id IN %s AND id IN %s
            """
            
            self.env.cr.execute(query, (codes_tuple, ids_tuple))
            result = self.env.cr.fetchall()
            
            ids = [row[0] for row in result]
            
        objects = objects.browse(ids)
        data_list = []     
        for obj in objects:
            data_dict = {                
                'id': obj.product_id.id,            
                'Código': obj.default_code,
                'Medida': obj.tire_measure_id.name if obj.tire_measure_id else None,
                'Capas': obj.layer_id.name if obj.layer_id else None,
                'Vel': obj.speed_id.name if obj.speed_id else None,
                'Carga': obj.index_of_load_id.name if obj.index_of_load_id else None,
                'Modelo': obj.model_id.name if obj.model_id else None,
                'Marca': obj.brand_id.name if obj.brand_id else None,
                'Tipo': obj.product_id.type_id.name or None,
                'Seg': obj.product_id.segment_id.name or None,
                'Tier': obj.tier_id.name if obj.tier_id else None,
                'Inv.': obj.inventario_str,
                'Cantidad Disponible': obj.available,
                'Trans.': (obj.transito_str + obj.backorder_str) or None,
                'Arribo': obj.fecha_aprox or None,
                'DOT': obj.product_id.product_dot_range or 'N/A' or None,  
                }
            if self.partner_id:
                data_dict.update({'Volumen': self.calcular_porcentaje(self.calcular_porcentaje(obj.volumen, 16, 'suma'), self.partner_id.volume_profile.percent, 'resta'),})
            else:
                data_dict.update({'Volumen': self.calcular_porcentaje(self.calcular_porcentaje(obj.volumen, 16, 'suma'), self.volume_profile.percent, 'resta'),})
            data_list.append(data_dict)            
            data_dict.update({
                'Promoción': obj.promocion * 1.16,
                'Outlet': obj.outlet * 1.16,
                'Mejor Condición': "",
                'Descuento Potencial': "",
                'Financiero': "",
                'Pedido': "",
                'Total': "",
            })
        return data_list
    

    def set_frames(self, sheet):        
        img_path = "/mnt/extra-addons/inv_promo/wizard/models/Encabezado.png"    
        img = Image(img_path)
        img.height = 58.5  # Establece la altura en píxeles calculados
        img.width = 1100   # Establece el ancho en píxeles calculados                       
        # Insertar la imagen en la hoja
        sheet.add_image(img, 'A1')

    def get_xlsx_report(self, partner_id, return_base_64=False):
        
        include_promo = True
        
        objects = self.env['inv_promo.report'].sudo().insert_data()
        con_iva = self.get_dict_data(objects,'Precios Con Iva',partner_id)        
        #precio_especial = self.get_dict_data(objects,'P. GOODYEAR')
        promo_bridgestone = self.get_dict_data(objects, 'P. BRIDGESTONE', partner_id)
        promocion23 = self.get_dict_data(objects, 'Onyx', partner_id)
        promo_tier4 = self.get_dict_data(objects, 'Tier 4', partner_id)
        wb = Workbook()
        sheet1 = wb.active
        sheet1.title = "Precios Con Iva"

        financial,logistic = self.get_profile_data(partner_id)
        #bs_percent = ['0%','4%','6%','8%','10%','12%']
         
        combo_data_sheet1 = [
            {'cell_ref': 'T13', 'values': logistic},
            {'cell_ref': 'V13', 'values': financial},
            #{'cell_ref': 'U13', 'values': bs_percent}
        ]
        
        self.set_combos(sheet1, combo_data_sheet1)
      
        #####Fill Tables
        num_rows_1, num_cols_1 = self.insert_table_sheet1_3(sheet1, con_iva, "TablaDatos1")

        self.set_frames(sheet1)
        
        sheet1.sheet_properties.tabColor = WHITE

        self.set_white_fill(sheet1, 'A1', num_rows_1, num_cols_1)
        
         # Definir los datos para las celdas
        data_for_sheet1 = [
            {'cell_ref': 'T4:W4', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': BLACK, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0', 'value': "Inventario + Promociones Diciembre 2024", 'data_type': 'string'},
            {'cell_ref': 'T5:W6', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': True, 'fill_color': WHITE, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': True, 'num_format': '#,##0', 'value': "Consulte en su correo, por whatsapp y/o con su asesor, todas las Promociones del Mes.", 'data_type': 'string'},
            {'cell_ref': 'W7', 'font_name': 'Calibri', 'font_size': 11, 'font_color': '808080', 'bold': True, 'fill_color': WHITE, 'border': 'none', 'align': 'left', 'top_align': 'top', 'wrap_text': False, 'num_format': '#,##0.00', 'value': f"V{self.partner_id.volume_profile.name}F{self.partner_id.financial_profile.letter}L{self.partner_id.logistic_profile.letter}", 'data_type': 'string'},
            {'cell_ref': 'A5:H5', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': True, 'fill_color': WHITE, 'border': 'thin', 'align': 'left', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0', 'value': "Términos y Condiciones", 'data_type': 'string'},
            {'cell_ref': 'A6:H6', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': False, 'fill_color': 'FFFFCB', 'border': 'thin', 'align': 'left', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0', 'value': "Precios Outlet: Solo aplica descuento financiero.", 'data_type': 'string'},
            {'cell_ref': 'A7:H7', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': False, 'fill_color': 'DDE5F2', 'border': 'thin', 'align': 'left', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0', 'value': "Códigos en Promoción, consulte segunda hoja.", 'data_type': 'string'},
            {'cell_ref': 'A8:H8', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': False, 'fill_color': 'FECCCB', 'border': 'thin', 'align': 'left', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0', 'value': "Códigos que suman en la promoción Bridgestone, consulte tercera hoja.", 'data_type': 'string'},
            {'cell_ref': 'A9:H9', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': False, 'fill_color': 'D9D9D9', 'border': 'thin', 'align': 'left', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0', 'value': "Códigos Tier 4 en Promoción, consulte cuarta hoja.", 'data_type': 'string'},
            {'cell_ref': 'A10:H10', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': True, 'fill_color': WHITE, 'border': 'thin', 'align': 'left', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0', 'value': "Una vez salida la mercancía, no se aceptan devoluciones.", 'data_type': 'string'},
            {'cell_ref': 'U13', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thick', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '0%', 'value': "0%"},
            {'cell_ref': 'V13', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': '00B04F', 'border': 'thick', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '0%', 'value': "0%"},
            {'cell_ref': 'T13', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': True, 'fill_color': 'F2F2F2', 'border': 'thick', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '0%', 'value': "0%"},
            {'cell_ref': 'W13', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0', 'value': f"=SUM(W15:W{num_rows_1})", 'data_type': 'string'},
            {'cell_ref': 'T11', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': True, 'fill_color': 'F2F2F2', 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0', 'value': "Descuento Logístico", 'data_type': 'string'},
            {'cell_ref': 'T12', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': True, 'fill_color': 'F2F2F2', 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0', 'value': f"""=IF(T13 = 0%, "Sin descuento", IF(T13 = 1%, "Min 80pz",  IF(T13 = 2%, "Min 250pz", IF(T13 = 4%,"Min 500pz","Sin descuento"))))""", 'data_type': 'string'},
            {'cell_ref': 'A14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
            {'cell_ref': 'B14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
            {'cell_ref': 'C14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
            {'cell_ref': 'D14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
            {'cell_ref': 'E14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
            {'cell_ref': 'F14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
            {'cell_ref': 'G14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
            {'cell_ref': 'H14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
            {'cell_ref': 'I14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
            {'cell_ref': 'J14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
            {'cell_ref': 'K14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
            {'cell_ref': 'L14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
            {'cell_ref': 'M14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
            {'cell_ref': 'N14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
            {'cell_ref': 'O14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
            {'cell_ref': 'P14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
            {'cell_ref': 'Q14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
            {'cell_ref': 'R14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
            {'cell_ref': 'S14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
            {'cell_ref': 'T14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
            {'cell_ref': 'U14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
            {'cell_ref': 'V14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': '00B04F', 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
            {'cell_ref': 'W14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': '595959', 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
            {'cell_ref': 'X14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
         ]
        
        #Establecer formatos
        self.format_column(sheet1,'B','string')
        self.format_column(sheet1,'C','string')
        self.format_column(sheet1,'D','string')
        self.format_column(sheet1,'E','string')
        self.format_column(sheet1,'F','string')
        self.format_column(sheet1,'G','string')
        self.format_column(sheet1,'H','string')
        self.format_column(sheet1,'I','string')
        self.format_column(sheet1,'J','string')
        self.format_column(sheet1,'K','string')
        self.format_column(sheet1,'L','string')
        self.format_column(sheet1,'M','number')
        self.format_column(sheet1,'N','number')
        self.format_column(sheet1,'O','date')
        self.format_column(sheet1,'P','string')
        self.format_column(sheet1,'Q','currency')
        self.format_column(sheet1,'R','currency')
        self.format_column(sheet1,'S','currency')
        self.format_column(sheet1,'T','currency')
        self.format_column(sheet1,'U','currency')
        self.format_column(sheet1,'V','currency')
        self.format_column(sheet1,'W','number')
        self.format_column(sheet1,'X','currency')
        self.format_column(sheet1,'Z','currency')
        
        columna_a_ocultar_sheet1 = ['A', 'M', 'Q', 'R', 'S', 'U', 'X']
        
        for columna_sheet1 in columna_a_ocultar_sheet1:
            sheet1.column_dimensions[columna_sheet1].hidden = True
        
        # Formatear las celdas para la hoja 1
        for data in data_for_sheet1:
            self.format_cell(sheet1, **data)
            
        self.color_cells_based_on_condition(sheet1)
        sheet1.sheet_view.showGridLines = False
        sheet1.freeze_panes = 'A15'
        
        if include_promo:
            sheet4 = wb.create_sheet(title="Onyx")
            sheet3 = wb.create_sheet(title="P. BRIDGESTONE")
            sheet5 = wb.create_sheet(title="Tier 4")
            
            combo_data_sheet2 = [
            {'cell_ref': 'U13', 'values': financial},    
            ]  
            
            self.set_combos(sheet3, combo_data_sheet1)  
            
            num_rows_3, num_cols_3 = self.insert_table_sheet1_3(sheet3, promo_bridgestone, "TablaDatos3")
            num_rows_3, num_cols_3 = self.insert_table_sheet4(sheet4, promocion23, "TablaDatos4")
            num_rows_3, num_cols_3 = self.insert_table_sheet5(sheet5, promo_tier4, "TablaDatos5")

            self.set_frames(sheet3)
            self.set_frames(sheet4)
            self.set_frames(sheet5)
            
            sheet3.sheet_properties.tabColor = 'E6B8B8'
            sheet4.sheet_properties.tabColor = '01B0F1'
            sheet5.sheet_properties.tabColor = '808080'
            
            self.set_white_fill(sheet3, 'A1', num_rows_3, num_cols_3)
            self.set_white_fill(sheet4, 'A1', num_rows_3, num_cols_3)
            self.set_white_fill(sheet5, 'A1', num_rows_3, num_cols_3)

            data_for_sheet3 = [
                {'cell_ref': 'E1', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': True, 'fill_color': WHITE, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'value': "", 'data_type': 'string'},
                {'cell_ref': 'T13', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': True, 'fill_color': 'BAB6B5', 'border': 'thick', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '0%', 'value': "0%"},
                {'cell_ref': 'V13', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thick', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '0%', 'value': "0%"},
                {'cell_ref': 'T12', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': True, 'fill_color': 'BAB6B5', 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0', 'value': f"""=IF(T13 = 0%, "Sin descuento", IF(T13 = 1%, "Min 80pz",  IF(T13 = 2%, "Min 250pz", IF(T13 = 4%,"Min 500pz","Sin descuento"))))""", 'data_type': 'string'},
                {'cell_ref': 'W13', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thick', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0', 'value': f"=SUM(W15:W{num_rows_3})", 'data_type': 'string'},
                {'cell_ref': 'A14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'B14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'C14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'D14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'E14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'F14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'G14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'H14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'I14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'J14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'K14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'L14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'M14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'N14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'O14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'P14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'Q14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'R14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'S14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'T14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'U14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'V14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'W14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'Y14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
            ]
            
            data_for_sheet4 = [
                {'cell_ref': 'A5:H5', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': BLACK, 'border': 'thin', 'align': 'left', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0', 'value': "Onyx", 'data_type': 'string'},
                {'cell_ref': 'A6:H7', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': False, 'fill_color': '00B0F0', 'border': 'thin', 'align': 'left', 'top_align': 'center', 'wrap_text': True, 'num_format': '#,##0', 'value': "EL PRECIO MOSTRADO DE PROMOCIÓN NO SE VE REFLEJADO EN SU FACTURA ya que solo es el estimado al aplicar la Nota de Crédito con la bonificación alcanzada al finalizar el mes.", 'data_type': 'string'},
                {'cell_ref': 'A8:H8', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': False, 'fill_color': 'FFFFCB', 'border': 'thin', 'align': 'left', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0', 'value': "Precios Outlet: Solo aplica descuento financiero.", 'data_type': 'string'},
                {'cell_ref': 'A9:H9', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': True, 'fill_color': WHITE, 'border': 'thin', 'align': 'left', 'top_align': 'center', 'wrap_text': True, 'num_format': '#,##0', 'value': "Una vez salida la mercancía, no se aceptan devoluciones.", 'data_type': 'string'},
                {'cell_ref': 'A10:H10', 'font_name': 'Calibri', 'font_size': 11, 'font_color': '00B050', 'bold': True, 'fill_color': WHITE, 'border': 'thin', 'align': 'left', 'top_align': 'center', 'wrap_text': True, 'num_format': '#,##0', 'value': "Aplica descuento financiero según su fecha de pago.", 'data_type': 'string'}, 
                {'cell_ref': 'E1', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': True, 'fill_color': WHITE, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'value': "", 'data_type': 'string'},
                {'cell_ref': 'V11:X11', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': '01B0F1', 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': True, 'num_format': '#,##0', 'value': "PROMOCIÓN DICIEMBRE", 'data_type': 'string'},
                {'cell_ref': 'Y11', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': '01B0F1', 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0', 'value': "ACUMULADO", 'data_type': 'string'}, 
                {'cell_ref': 'V12', 'font_name': 'Calibri', 'font_size': 11, 'font_color': '538DD5', 'bold': True, 'fill_color': 'C5D9F1', 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '0%', 'value': "4%"},
                {'cell_ref': 'W12', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': '8DB4E2', 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '0%', 'value': "8%"},
                {'cell_ref': 'X12', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': 'C5D9F1', 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '0%', 'value': "10%"},
                {'cell_ref': 'Y12:Y13', 'font_name': 'Calibri', 'font_size': 11, 'font_color': '0070C0', 'bold': True, 'fill_color': 'DCE6F1', 'border': 'thick', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0', 'value': f"=SUM(Y15:Y{num_rows_3})", 'data_type': 'string'}, 
                {'cell_ref': 'V13:X13', 'font_name': 'Calibri', 'font_size': 11, 'font_color': '538DD5', 'bold': True, 'fill_color': 'C5D9F1', 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '0%', 'value': "En la compra de más de"},
                {'cell_ref': 'A14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'B14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'C14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'D14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'E14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'F14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'G14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'H14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'I14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'J14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'K14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'L14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'M14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'N14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'O14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'P14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'Q14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'R14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'S14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'T14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'U14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'V14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': '538DD5', 'bold': True, 'fill_color': 'C5D9F1', 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'W14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': '8DB4E2', 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'X14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': 'C5D9F1', 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'Y14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
            ]
            
            data_for_sheet5 = [
                {'cell_ref': 'A5:H5', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': BLACK, 'border': 'thin', 'align': 'left', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0', 'value': "Tier 4", 'data_type': 'string'},
                {'cell_ref': 'A6:H7', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': False, 'fill_color': 'BFBFBF', 'border': 'thin', 'align': 'left', 'top_align': 'center', 'wrap_text': True, 'num_format': '#,##0', 'value': "EL PRECIO MOSTRADO DE PROMOCIÓN NO SE VE REFLEJADO EN SU FACTURA ya que solo es el estimado al aplicar la Nota de Crédito con la bonificación alcanzada al finalizar el mes.", 'data_type': 'string'},
                {'cell_ref': 'A8:H8', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': False, 'fill_color': 'FFFFCB', 'border': 'thin', 'align': 'left', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0', 'value': "Precios Outlet: Solo aplica descuento financiero.", 'data_type': 'string'},
                {'cell_ref': 'A9:H9', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': True, 'fill_color': WHITE, 'border': 'thin', 'align': 'left', 'top_align': 'center', 'wrap_text': True, 'num_format': '#,##0', 'value': "Una vez salida la mercancía, no se aceptan devoluciones.", 'data_type': 'string'},
                {'cell_ref': 'A10:H10', 'font_name': 'Calibri', 'font_size': 11, 'font_color': '00B050', 'bold': True, 'fill_color': WHITE, 'border': 'thin', 'align': 'left', 'top_align': 'center', 'wrap_text': True, 'num_format': '#,##0', 'value': "Aplica descuento financiero según su fecha de pago.", 'data_type': 'string'}, 
                {'cell_ref': 'E1', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': True, 'fill_color': WHITE, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'value': "", 'data_type': 'string'},
                {'cell_ref': 'V11:X11', 'font_name': 'Calibri', 'font_size': 11, 'font_color': '000000', 'bold': True, 'fill_color': 'BFBFBF', 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': True, 'num_format': '#,##0', 'value': "PROMOCIÓN DICIEMBRE", 'data_type': 'string'},
                {'cell_ref': 'Y11', 'font_name': 'Calibri', 'font_size': 11, 'font_color': '000000', 'bold': True, 'fill_color': 'BFBFBF', 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0', 'value': "LLANTAS ACUMULADAS", 'data_type': 'string'}, 
                {'cell_ref': 'V12', 'font_name': 'Calibri', 'font_size': 11, 'font_color': '808080', 'bold': True, 'fill_color': 'F2F2F2', 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '0%', 'value': "3%"},
                {'cell_ref': 'W12', 'font_name': 'Calibri', 'font_size': 11, 'font_color': '808080', 'bold': True, 'fill_color': 'BFBFBF', 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '0%', 'value': "6%"},
                {'cell_ref': 'X12', 'font_name': 'Calibri', 'font_size': 11, 'font_color': '808080', 'bold': True, 'fill_color': 'F2F2F2', 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '0%', 'value': "8%"},
                {'cell_ref': 'V13:X13', 'font_name': 'Calibri', 'font_size': 11, 'font_color': '808080', 'bold': True, 'fill_color': 'F2F2F2', 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '0%', 'value': "En la compra de más de"},
                {'cell_ref': 'Y12:Y13', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': '595959', 'border': 'thick', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0', 'value': f"=SUM(Y15:Y{num_rows_3})", 'data_type': 'string'},
                {'cell_ref': 'A14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'B14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'C14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'D14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'E14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'F14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'G14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'H14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'I14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'J14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'K14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'L14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'M14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'N14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'O14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'P14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'Q14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'R14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'S14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'T14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'U14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'V14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': '808080', 'bold': True, 'fill_color': 'F2F2F2', 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'W14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': '808080', 'bold': True, 'fill_color': 'BFBFBF', 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'X14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': '808080', 'bold': True, 'fill_color': 'F2F2F2', 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
                {'cell_ref': 'Y14', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': DARK_GRAY, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'data_type': 'string'},
            ]
            
            promobrid = [
                {'cell_ref': 'A5:H5', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': BLACK, 'border': 'thin', 'align': 'left', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0', 'value': "PROMOCIÓN BRIDGESTONE", 'data_type': 'string'},
                {'cell_ref': 'A6:H7', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': False, 'fill_color': 'FFCCCC', 'border': 'thin', 'align': 'left', 'top_align': 'center', 'wrap_text': True, 'num_format': '#,##0', 'value': "EL PRECIO MOSTRADO NO SE VE REFLEJADO EN SU FACTURA ya que solo es el estimado al aplicar la nota de crédito con la bonificación alcanzada al finalizar el mes.", 'data_type': 'string'},
                {'cell_ref': 'A8:H8', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': False, 'fill_color': 'FFFFCB', 'border': 'thin', 'align': 'left', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0', 'value': "Precios Outlet: Solo aplica descuento financiero.", 'data_type': 'string'},
                {'cell_ref': 'A9:H9', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': True, 'fill_color': WHITE, 'border': 'thin', 'align': 'left', 'top_align': 'center', 'wrap_text': True, 'num_format': '#,##0', 'value': "Una vez salida la mercancía, no se aceptan devoluciones.", 'data_type': 'string'},
                {'cell_ref': 'A10:H10', 'font_name': 'Calibri', 'font_size': 11, 'font_color': '00B050', 'bold': True, 'fill_color': WHITE, 'border': 'thin', 'align': 'left', 'top_align': 'center', 'wrap_text': True, 'num_format': '#,##0', 'value': "Aplica descuento financiero según su fecha de pago.", 'data_type': 'string'}, 
                {'cell_ref': 'T5:W5', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': RED, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'value': "PROMOCIÓN BRIDGESTONE", 'data_type': 'string'},
                {'cell_ref': 'T6', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': BLACK, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'value': "Min", 'data_type': 'string'},
                {'cell_ref': 'U6:V6', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': BLACK, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'value': "Max", 'data_type': 'string'},
                {'cell_ref': 'W6', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': BLACK, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'value': "Bonificación", 'data_type': 'string'},
                {'cell_ref': 'T7', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': True, 'fill_color': WHITE, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'value': 46400, 'data_type': 'string'},
                {'cell_ref': 'U7:V7', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': True, 'fill_color': WHITE, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'value': 115999, 'data_type': 'string'},
                {'cell_ref': 'W7', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': True, 'fill_color': WHITE, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'value': "4%", 'data_type': 'string'},
                {'cell_ref': 'T8', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': True, 'fill_color': WHITE, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'value': 116000, 'data_type': 'string'},
                {'cell_ref': 'U8:V8', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': True, 'fill_color': WHITE, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'value': 347999, 'data_type': 'string'},
                {'cell_ref': 'W8', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': True, 'fill_color': WHITE, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'value': "6%", 'data_type': 'string'},
                {'cell_ref': 'T9', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': True, 'fill_color': WHITE, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'value': 348000, 'data_type': 'string'},
                {'cell_ref': 'U9:V9', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': True, 'fill_color': WHITE, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'value': 927999, 'data_type': 'string'},
                {'cell_ref': 'W9', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': True, 'fill_color': WHITE, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'value': "8%", 'data_type': 'string'},
                {'cell_ref': 'T10', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': True, 'fill_color': WHITE, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'value': 928000, 'data_type': 'string'},
                {'cell_ref': 'U10:V10', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': True, 'fill_color': WHITE, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'value': "o más", 'data_type': 'string'},
                {'cell_ref': 'W10', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': True, 'fill_color': WHITE, 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'value': "10%", 'data_type': 'string'},
                {'cell_ref': 'W11', 'font_name': 'Calibri', 'font_size': 11, 'font_color': WHITE, 'bold': True, 'fill_color': 'FF0000', 'border': 'thin', 'align': 'center', 'top_align': 'center', 'wrap_text': False, 'num_format': '#,##0.00', 'value': "ACUMULADO", 'data_type': 'string'},
                {'cell_ref': 'W12', 'font_name': 'Calibri', 'font_size': 11, 'font_color': '404040', 'bold': True, 'fill_color': 'FFCCCC', 'border': 'thin', 'align': 'center', 'top_align': 'top', 'wrap_text': False, 'num_format': '$#,##0.00', 'value': f"=SUM(X15:X{num_rows_3})", 'data_type': 'currency'},
                {'cell_ref': 'U13', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': True, 'fill_color': 'BAB6B5', 'border': 'thick', 'align': 'center', 'top_align': 'top', 'wrap_text': False, 'num_format': '#,##0.00', 'value': f"""=IF(AND(W12 >= T7, W12 <= U7), W7, IF(AND(W12 >= T8, W12 <= U8), W8, IF(AND(W12 >= T9, W12 <= U9), W9, IF(W12 >= T10, W10, "0%"))))""", 'data_type': 'string'},
                {'cell_ref': 'U12', 'font_name': 'Calibri', 'font_size': 11, 'font_color': BLACK, 'bold': True, 'fill_color': 'BAB6B5', 'border': 'thick', 'align': 'center', 'top_align': 'top', 'wrap_text': False, 'num_format': '#,##0.00', 'value': f"""=IF(U13 = 0%, "Sin descuento potencial", IF(U13 = 4%, U2 & " " & TEXT(U3,"$#,##0.00"), IF(U13 = 6%,  U2 & " " & TEXT(U4,"$#,##0.00"), IF(U13 = 8%,  U2 & " " & TEXT(U5,"$#,##0.00"), IF(U13 = 10%,  U2 & " " & TEXT(U6,"$#,##0.00"), IF(U13 = 12%,  U2 & " " & TEXT(U7,"$#,##0.00"), "Sin descuento potencial"))))))""", 'data_type': 'string'}
            ]

            # Formatear las celdas para la hoja 3
            for data in data_for_sheet3:
                self.format_cell(sheet3, **data)
            for data in promobrid:
                self.format_cell(sheet3, **data)
                
            for data in data_for_sheet4:
                self.format_cell(sheet4, **data)
                
            for data in data_for_sheet5:
                self.format_cell(sheet5, **data)
            
            # Lista de hojas a formatear
            sheets = [sheet3, sheet4, sheet5]

            # Diccionario que define las columnas y sus formatos
            formats = {
                'string': ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'P'],
                'number': ['M', 'N', 'W'],
                'date': ['O'],
                'currency': ['Q', 'R', 'S', 'T', 'U', 'V', 'X']
            }
            
            formats3 = {
                'string': ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'P', 'R'],
                'number': ['M', 'N', 'Q'],
                'date': ['O'],
                'currency': ['S', 'T', 'U', 'V', 'W', 'X', 'Y']
            }
            #-------- Ocultar columnas ----------------------
            # Seleccionar la hoja que deseas ocultar
            # sheet = wb['P. BRIDGESTONE']
            # Para ocultar la hoja:
            # sheet.sheet_state = 'hidden'
            
            columna_a_ocultar_sheet3 = ['A', 'M', 'Q', 'R', 'S', 'V', 'X']
            columna_a_ocultar_sheet4 = ['A', 'M', 'N', 'O', 'Q', 'R', 'S', 'U', 'Z']

            # Itera sobre cada hoja
            for sheet in sheets:
                if sheet == sheet4 or sheet == sheet5:
                    for format_type, columns in formats3.items():
                        for column in columns:
                            self.format_column(sheet, column, format_type)
                            #-----------Ocultar cuadricula---------------------------
                            if sheet == sheet4:
                                self.color_cells_based_on_condition_sheet4(sheet)
                            sheet.sheet_view.showGridLines = False
                            sheet.freeze_panes = 'A15'
                    for columna_sheet4 in columna_a_ocultar_sheet4:
                        sheet.column_dimensions[columna_sheet4].hidden = True
                else:
                    # Itera sobre cada tipo de formato y sus respectivas columnas
                    for format_type, columns in formats.items():
                        for column in columns:
                            self.format_column(sheet, column, format_type)
                            #-----------Ocultar cuadricula---------------------------
                            self.color_cells_based_on_condition_sheet3(sheet)
                            sheet.sheet_view.showGridLines = False
                            sheet.freeze_panes = 'A15'
                    for columna_sheet3 in columna_a_ocultar_sheet3:
                        sheet.column_dimensions[columna_sheet3].hidden = True
                        
        # Ajustar el ancho de las columnas en todas las hojas basándose en la fila 1 (por ejemplo)
        row_num = 14
        desired_height = 40  # Altura deseada en puntos
        
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            ws.row_dimensions[row_num].height = desired_height
            self.auto_adjust_column_widths(ws, row_num)
            
        if 'Tier 4' in wb.sheetnames:
            del wb['Tier 4']
            
        propiedades = wb.properties
        # Crear un buffer en memoria
        excel_bytes_io = io.BytesIO()

        # Guardar el archivo en el buffer
        wb.save(excel_bytes_io)

        # Obtener los bytes del archivo y codificarlo en base64
        excel_bytes_io.seek(0)  # Rewind the buffer to the beginning before reading it
        excel_base64 = base64.b64encode(excel_bytes_io.read()).decode("utf-8")

        # Si deseas devolver el archivo como base64, puedes hacerlo:
        if return_base_64:
            return excel_base64
        
        self.write({'file_data': excel_base64})
        
        action = {
            'name': 'Lista de Precios',
            'type': 'ir.actions.act_url',
            'url': f'/web/content/?model=inv_promo.lista_precios_wizard&id={self.id}&field=file_data&download=true&filename={self.partner_id.name or "Lista de Precios M"}.xlsx',
            'target': '_blank',
            }
        return action
    
    def auto_adjust_column_widths(self, ws, row_num):
        row = ws[row_num]
        for cell in row:
            column_index = cell.column
            column_letter = get_column_letter(column_index)
            max_length = len(str(cell.value)) if cell.value else 0
            adjusted_width = max_length + 5
            ws.column_dimensions[column_letter].width = adjusted_width

    def download_report(self):
        if self.partner_id:
            return self.sudo().get_xlsx_report(self.partner_id)
        elif not self.partner_id and self.volume_profile and self.financial_profile and self.logistic_profile:
            return self.sudo().get_xlsx_report(self.partner_id)